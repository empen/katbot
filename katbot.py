#katbot?
# import data from "inventory location, verathon tab"
# spit out verathon inbound summary

# Ignore any VAN, HOLD, QC, FGUSRL STV

# FGURLS - Lopro, mac, miller - (exception NS) - (35 days) from birthday
# GVL's, pedo stylets, single use stylets, 0270-1093 - (23 days) from birthday
# reuse stylet - 14 days
# non strile lopro/mac - 14 days
# spot out based on week
#
from datetime import date
import datetime
today = date.today()
current_week = datetime.date(date.today().year, date.today().month, date.today().day).isocalendar()[1]
# print(current_week)
# print("Running this hullabaloo for: ",today)
from openpyxl import Workbook, load_workbook

verathon_inventory = load_workbook(filename = 'Files\Inventory Location.xlsx')
weekly_invin = verathon_inventory["Verathon"]


#lists for Verathon Export
su_lopro_1 = ["SU Lopro 1", "0270-0876"]
su_lopro_2 = ["SU Lopro 2", "0270-0877"]
su_lopro_25 = ["SU Lopro 2.5", "0270-0962"]
su_lopro_3 = ["SU Lopro3 - Sterile", "0270-0938"]
su_lopro_4 = ["SU Lopro4 - Sterile", "0270-0939"]
ns_lopro_3 = ["SU Lopro3 - Non-Sterile", "0270-1078"]
ns_lopro_4 = ["SU Lopro4 - Non-Sterile", "0270-1079"]
miller_0 = ["Miller 0", "0270-0966"]
miller_1 = ["Miller 1", "0270-0967"]
mac_3 = ["DVM S3 - Sterile", "0270-0932"]
mac_4 = ["DVM S4 - Sterile", "0270-0933"]
ns_mac_3 = ["DVM S3 - Non-Sterile", "0270-1080"]
ns_mac_4 = ["DVM S4 - Non-Sterile", "0270-1081"]
gvl_0 = ["GVL 0", "0270-0679"]
gvl_1 = ["GVL 1", "0270-0428"]
gvl_2 = ["GVL 2", "0270-0419"]
gvl_25 = ["GVL2.5", "0270-0797"]
gvl_3 = ["GVL Stat 3", "0270-0626"]
gvl_4 = ["GVL Stat 4", "0270-0628"]
ped_stylet = ["Pediatric Stylet", "0270-0916"]

# if part number matches "VER1/0270-0967" and location is NOT VAN, HOLD, QC, FGUSRL STV + 35 days to date. add to nearest week

#6 week chunk of time separated
week_1_35 = today - datetime.timedelta(days=7)
week_2_35 = today
week_3_35 = today + datetime.timedelta(days=14)
week_4_35 = today + datetime.timedelta(days=21)
week_5_35 = today + datetime.timedelta(days=28)
week_6_35 = today + datetime.timedelta(days=35)

week_1_23 = today - datetime.timedelta(days=7)
week_2_23 = today
week_3_23 = today + datetime.timedelta(days=14)
week_4_23 = today + datetime.timedelta(days=21)
week_5_23 = today + datetime.timedelta(days=28)
week_6_23 = today + datetime.timedelta(days=35)
#Find first day of each week

rows = weekly_invin.iter_rows(min_row=1, max_row=250, min_col=1, max_col=7)
# Sums of each type of verathon product for each week
# PED SUMS
miller_0_total_1 = 0
miller_0_total_2 = 0
miller_0_total_3 = 0
miller_0_total_4 = 0
miller_0_total_5 = 0
miller_0_total_6 = 0
miller_1_total_1 = 0
miller_1_total_2 = 0
miller_1_total_3 = 0
miller_1_total_4 = 0
miller_1_total_5 = 0
miller_1_total_6 = 0
s1_total_1 = 0
s1_total_2 = 0
s1_total_3 = 0
s1_total_4 = 0
s1_total_5 = 0
s1_total_6 = 0
s2_total_1 = 0
s2_total_2 = 0
s2_total_3 = 0
s2_total_4 = 0
s2_total_5 = 0
s2_total_6 = 0
s25_total_1 = 0
s25_total_2 = 0
s25_total_3 = 0
s25_total_4 = 0
s25_total_5 = 0
s25_total_6 = 0

#ADULT SUMS
s3_total_1 = 0
s3_total_2 = 0
s3_total_3 = 0
s3_total_4 = 0
s3_total_5 = 0
s3_total_6 = 0
s3_ns_total_1 = 0
s3_ns_total_2 = 0
s3_ns_total_3 = 0
s3_ns_total_4 = 0
s3_ns_total_5 = 0
s3_ns_total_6 = 0
s4_total_1 = 0
s4_total_2 = 0
s4_total_3 = 0
s4_total_4 = 0
s4_total_5 = 0
s4_total_6 = 0
s4_ns_total_1 = 0
s4_ns_total_2 = 0
s4_ns_total_3 = 0
s4_ns_total_4 = 0
s4_ns_total_5 = 0
s4_ns_total_6 = 0
mac_3_total_1 = 0
mac_3_total_2 = 0
mac_3_total_3 = 0
mac_3_total_4 = 0
mac_3_total_5 = 0
mac_3_total_6 = 0
mac_3_ns_total_1 = 0
mac_3_ns_total_2 = 0
mac_3_ns_total_3 = 0
mac_3_ns_total_4 = 0
mac_3_ns_total_5 = 0
mac_3_ns_total_6 = 0
mac_4_total_1 = 0
mac_4_total_2 = 0
mac_4_total_3 = 0
mac_4_total_4 = 0
mac_4_total_5 = 0
mac_4_total_6 = 0
mac_4_ns_total_1 = 0
mac_4_ns_total_2 = 0
mac_4_ns_total_3 = 0
mac_4_ns_total_4 = 0
mac_4_ns_total_5 = 0
mac_4_ns_total_6 = 0

# GVL Totals
gvl_3_total_1 = 0
gvl_3_total_2 = 0
gvl_3_total_3 = 0
gvl_3_total_4 = 0
gvl_3_total_5 = 0
gvl_3_total_6 = 0
gvl_4_total_1 = 0
gvl_4_total_2 = 0
gvl_4_total_3 = 0
gvl_4_total_4 = 0
gvl_4_total_5 = 0
gvl_4_total_6 = 0

for a,b,c,d,e,f,g in rows:
    if isinstance(g.value, datetime.datetime):
        second_coming_35 = g.value.date() + datetime.timedelta(days=35)
        second_coming_23 = g.value.date() + datetime.timedelta(days=23)
    #SU LOPRO 1 WORK
    if a.value == "VER1/0270-0876" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_1_35 and second_coming_35 < week_2_35:
        s1_total_1 += f.value
    if a.value == "VER1/0270-0876" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_2_35 and second_coming_35 < week_3_35:
        s1_total_2 += f.value
    if a.value == "VER1/0270-0876" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_3_35 and second_coming_35 < week_4_35:
        s1_total_3 += f.value
    if a.value == "VER1/0270-0876" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_4_35 and second_coming_35 < week_5_35:
        s1_total_4 += f.value
    if a.value == "VER1/0270-0876" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_5_35 and second_coming_35 < week_6_35:
        s1_total_5 += f.value
    if a.value == "VER1/0270-0876" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_6_35:
        s1_total_6 += f.value

    #SU LOPRO 2 WORK
    if a.value == "VER1/0270-0877" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_1_35 and second_coming_35 < week_2_35:
        s2_total_1 += f.value
    if a.value == "VER1/0270-0877" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_2_35 and second_coming_35 < week_3_35:
        s2_total_2 += f.value
    if a.value == "VER1/0270-0877" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_3_35 and second_coming_35 < week_4_35:
        s2_total_3 += f.value
    if a.value == "VER1/0270-0877" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_4_35 and second_coming_35 < week_5_35:
        s2_total_4 += f.value
    if a.value == "VER1/0270-0877" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_5_35 and second_coming_35 < week_6_35:
        s2_total_5 += f.value
    if a.value == "VER1/0270-0877" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_6_35:
        s2_total_6 += f.value

    #SU LOPRO 2.5 WORK
    if a.value == "VER1/0270-0962" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_1_35 and second_coming_35 < week_2_35:
        s25_total_1 += f.value
    if a.value == "VER1/0270-0962" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_2_35 and second_coming_35 < week_3_35:
        s25_total_2 += f.value
    if a.value == "VER1/0270-0962" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_3_35 and second_coming_35 < week_4_35:
        s25_total_3 += f.value
    if a.value == "VER1/0270-0962" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_4_35 and second_coming_35 < week_5_35:
        s25_total_4 += f.value
    if a.value == "VER1/0270-0962" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_5_35 and second_coming_35 < week_6_35:
        s25_total_5 += f.value
    if a.value == "VER1/0270-0962" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_6_35:
        s25_total_6 += f.value

    #SU MILLER 0 WORK
    if a.value == "VER1/0270-0966" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_1_35 and second_coming_35 < week_2_35:
        miller_0_total_1_total_1 += f.value
    if a.value == "VER1/0270-0966" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_2_35 and second_coming_35 < week_3_35:
        miller_0_total_2 += f.value
    if a.value == "VER1/0270-0966" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_3_35 and second_coming_35 < week_4_35:
        miller_0_total_3 += f.value
    if a.value == "VER1/0270-0966" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_4_35 and second_coming_35 < week_5_35:
        miller_0_total_4 += f.value
    if a.value == "VER1/0270-0966" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_5_35 and second_coming_35 < week_6_35:
        miller_0_total_5 += f.value
    if a.value == "VER1/0270-0966" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_6_35:
        miller_0_total_6 += f.value

    #SU MILLER 1 WORK
    if a.value == "VER1/0270-0967" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_1_35 and second_coming_35 < week_2_35:
        miller_1_total_1 += f.value
    if a.value == "VER1/0270-0967" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_2_35 and second_coming_35 < week_3_35:
        miller_1_total_2 += f.value
    if a.value == "VER1/0270-0967" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_3_35 and second_coming_35 < week_4_35:
        miller_1_total_3 += f.value
    if a.value == "VER1/0270-0967" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_4_35 and second_coming_35 < week_5_35:
        miller_1_total_4 += f.value
    if a.value == "VER1/0270-0967" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_5_35 and second_coming_35 < week_6_35:
        miller_1_total_5 += f.value
    if a.value == "VER1/0270-0967" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_6_35:
        miller_1_total_6 += f.value

    #SU LOPRO 3 WORK
    if a.value == "VER1/0270-0938" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_1_35 and second_coming_35 < week_2_35:
        s3_total_1 += f.value
    if a.value == "VER1/0270-0938" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_2_35 and second_coming_35 < week_3_35:
        s3_total_2 += f.value
    if a.value == "VER1/0270-0938" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_3_35 and second_coming_35 < week_4_35:
        s3_total_3 += f.value
    if a.value == "VER1/0270-0938" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_4_35 and second_coming_35 < week_5_35:
        s3_total_4 += f.value
    if a.value == "VER1/0270-0938" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_5_35 and second_coming_35 < week_6_35:
        s3_total_5 += f.value
    if a.value == "VER1/0270-0938" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_6_35:
        s3_total_6 += f.value

    #SU NON-STERILE LOPRO 3 WORK
    if a.value == "VER1/0270-1078" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_1_35 and second_coming_35 < week_2_35:
        s3_ns_total_1 += f.value
    if a.value == "VER1/0270-1078" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_2_35 and second_coming_35 < week_3_35:
        s3_ns_total_2 += f.value
    if a.value == "VER1/0270-1078" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_3_35 and second_coming_35 < week_4_35:
        s3_ns_total_3 += f.value
    if a.value == "VER1/0270-1078" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_4_35 and second_coming_35 < week_5_35:
        s3_ns_total_4 += f.value
    if a.value == "VER1/0270-1078" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_5_35 and second_coming_35 < week_6_35:
        s3_ns_total_5 += f.value
    if a.value == "VER1/0270-1078" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_6_35:
        s3_ns_total_6 += f.value

    #SU LOPRO 4 WORK
    if a.value == "VER1/0270-0939" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_1_35 and second_coming_35 < week_2_35:
        s4_total_1 += f.value
    if a.value == "VER1/0270-0939" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_2_35 and second_coming_35 < week_3_35:
        s4_total_2 += f.value
    if a.value == "VER1/0270-0939" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_3_35 and second_coming_35 < week_4_35:
        s4_total_3 += f.value
    if a.value == "VER1/0270-0939" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_4_35 and second_coming_35 < week_5_35:
        s4_total_4 += f.value
    if a.value == "VER1/0270-0939" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_5_35 and second_coming_35 < week_6_35:
        s4_total_5 += f.value
    if a.value == "VER1/0270-0939" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_6_35:
        s4_total_6 += f.value

    #SU NON STERILE LOPRO 4 WORK
    if a.value == "VER1/0270-1079" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_1_35 and second_coming_35 < week_2_35:
        s4_ns_total_1 += f.value
    if a.value == "VER1/0270-1079" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_2_35 and second_coming_35 < week_3_35:
        s4_ns_total_2 += f.value
    if a.value == "VER1/0270-1079" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_3_35 and second_coming_35 < week_4_35:
        s4_ns_total_3 += f.value
    if a.value == "VER1/0270-1079" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_4_35 and second_coming_35 < week_5_35:
        s4_ns_total_4 += f.value
    if a.value == "VER1/0270-1079" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_5_35 and second_coming_35 < week_6_35:
        s4_ns_total_5 += f.value
    if a.value == "VER1/0270-1079" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_6_35:
        s4_ns_total_6 += f.value

    #MAC 3 WORK
    if a.value == "VER1/0270-0932" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_1_35 and second_coming_35 < week_2_35:
        mac_3_total_1 += f.value
    if a.value == "VER1/0270-0932" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_2_35 and second_coming_35 < week_3_35:
        mac_3_total_2 += f.value
    if a.value == "VER1/0270-0932" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_3_35 and second_coming_35 < week_4_35:
        mac_3_total_3 += f.value
    if a.value == "VER1/0270-0932" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_4_35 and second_coming_35 < week_5_35:
        mac_3_total_4 += f.value
    if a.value == "VER1/0270-0932" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_5_35 and second_coming_35 < week_6_35:
        mac_3_total_5 += f.value
    if a.value == "VER1/0270-0932" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_6_35:
        mac_3_total_6 += f.value

    #MAC 3 NS WORK
    if a.value == "VER1/0270-1080" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_23 >= week_1_35 and second_coming_23 < week_2_35:
        mac_3_ns_total_1 += f.value
    if a.value == "VER1/0270-1080" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_23 >= week_2_35 and second_coming_23 < week_3_35:
        mac_3_ns_total_2 += f.value
    if a.value == "VER1/0270-1080" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_23 >= week_3_35 and second_coming_23 < week_4_35:
        mac_3_ns_total_3 += f.value
    if a.value == "VER1/0270-1080" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_23 >= week_4_35 and second_coming_23 < week_5_35:
        mac_3_ns_total_4 += f.value
    if a.value == "VER1/0270-1080" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_23 >= week_5_35 and second_coming_23 < week_6_35:
        mac_3_ns_total_5 += f.value
    if a.value == "VER1/0270-1080" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_23 >= week_6_35:
        mac_3_ns_total_6 += f.value

    #MAC 4 WORK
    if a.value == "VER1/0270-0933" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_1_35 and second_coming_35 < week_2_35:
        mac_4_total_1 += f.value
    if a.value == "VER1/0270-0933" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_2_35 and second_coming_35 < week_3_35:
        mac_4_total_2 += f.value
    if a.value == "VER1/0270-0933" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_3_35 and second_coming_35 < week_4_35:
        mac_4_total_3 += f.value
    if a.value == "VER1/0270-0933" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_4_35 and second_coming_35 < week_5_35:
        mac_4_total_4 += f.value
    if a.value == "VER1/0270-0933" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_5_35 and second_coming_35 < week_6_35:
        mac_4_total_5 += f.value
    if a.value == "VER1/0270-0933" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_6_35:
        mac_4_total_6 += f.value

#MAC 4 NS WORK
    if a.value == "VER1/0270-1081" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_23 >= week_1_35 and second_coming_23 < week_2_35:
        mac_4_ns_total_1 += f.value
    if a.value == "VER1/0270-1080" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_23 >= week_2_35 and second_coming_23 < week_3_35:
        mac_4_ns_total_2 += f.value
    if a.value == "VER1/0270-1080" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_23 >= week_3_35 and second_coming_23 < week_4_35:
        mac_4_ns_total_3 += f.value
    if a.value == "VER1/0270-1080" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_23 >= week_4_35 and second_coming_23 < week_5_35:
        mac_4_ns_total_4 += f.value
    if a.value == "VER1/0270-1080" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_23 >= week_5_35 and second_coming_23 < week_6_35:
        mac_4_ns_total_5 += f.value
    if a.value == "VER1/0270-1080" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_23 >= week_6_35:
        mac_4_ns_total_6 += f.value

    #GVL 3 WORK
    if a.value == "VER1/0270-0626" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_1_35 and second_coming_35 < week_2_35:
        gvl_3_total_1 += f.value
    if a.value == "VER1/0270-0626" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_2_35 and second_coming_35 < week_3_35:
        gvl_3_total_2 += f.value
    if a.value == "VER1/0270-0626" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_3_35 and second_coming_35 < week_4_35:
        gvl_3_total_3 += f.value
    if a.value == "VER1/0270-0626" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_4_35 and second_coming_35 < week_5_35:
        gvl_3_total_4 += f.value
    if a.value == "VER1/0270-0626" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_5_35 and second_coming_35 < week_6_35:
        gvl_3_total_5 += f.value
    if a.value == "VER1/0270-0626" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_6_35:
        gvl_3_total_6 += f.value

    #GVL 4 WORK
    if a.value == "VER1/0270-0628" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_1_35 and second_coming_35 < week_2_35:
        gvl_4_total_1 += f.value
    if a.value == "VER1/0270-0628" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_2_35 and second_coming_35 < week_3_35:
        gvl_4_total_2 += f.value
    if a.value == "VER1/0270-0628" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_3_35 and second_coming_35 < week_4_35:
        gvl_4_total_3 += f.value
    if a.value == "VER1/0270-0628" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_4_35 and second_coming_35 < week_5_35:
        gvl_4_total_4 += f.value
    if a.value == "VER1/0270-0628" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_5_35 and second_coming_35 < week_6_35:
        gvl_4_total_5 += f.value
    if a.value == "VER1/0270-0628" and d.value != "HOLD" and d.value != "QCSTV" and d.value != "FGRLS VAN" and d.value != "QC" and d.value != "HOLD VAN" and d.value != "FGURLS STV" and second_coming_35 >= week_6_35:
        gvl_4_total_6 += f.value
#PED SUMS
su_lopro_1.extend([s1_total_1, s1_total_2, s1_total_3, s1_total_4, s1_total_5, s1_total_6])
su_lopro_2.extend([s2_total_1, s2_total_2, s2_total_3, s2_total_4, s2_total_5, s2_total_6])
su_lopro_25.extend([s25_total_1, s25_total_2, s25_total_3, s25_total_4, s25_total_5, s25_total_6])
miller_0.extend([miller_0_total_1,miller_0_total_2,miller_0_total_3,miller_0_total_4,miller_0_total_5,miller_0_total_6])
miller_1.extend([miller_1_total_1,miller_1_total_2,miller_1_total_3,miller_1_total_4,miller_1_total_5,miller_1_total_6])

#ADULT SUMS
su_lopro_3.extend([s3_total_1, s3_total_2, s3_total_3, s3_total_4, s3_total_5, s3_total_6])
su_lopro_4.extend([s4_total_1, s4_total_2, s4_total_3, s4_total_4, s4_total_5, s4_total_6])
ns_lopro_3.extend([s3_ns_total_1,s3_ns_total_2,s3_ns_total_3,s3_ns_total_4,s3_ns_total_5,s3_ns_total_6,])
ns_lopro_4.extend([s4_ns_total_1,s4_ns_total_2,s4_ns_total_3,s4_ns_total_4,s4_ns_total_5,s4_ns_total_6,])
mac_3.extend([mac_3_total_1, mac_3_total_2, mac_3_total_3, mac_3_total_4,mac_3_total_5,mac_3_total_6])
ns_mac_3.extend([mac_3_ns_total_1, mac_3_ns_total_2, mac_3_ns_total_3, mac_3_ns_total_4, mac_3_ns_total_5, mac_3_ns_total_6])
mac_4.extend([mac_4_total_1, mac_4_total_2, mac_4_total_3, mac_4_total_4, mac_4_total_5, mac_4_total_6])
ns_mac_4.extend([mac_4_ns_total_1,mac_4_ns_total_2,mac_4_ns_total_3,mac_4_ns_total_4,mac_4_ns_total_5,mac_4_ns_total_6,])

#GVL ADULT SUMS
gvl_3.extend([gvl_3_total_1,gvl_3_total_2,gvl_3_total_3,gvl_3_total_4,gvl_3_total_5,gvl_3_total_6,])
gvl_4.extend([gvl_4_total_1,gvl_4_total_2,gvl_4_total_3,gvl_4_total_4,gvl_4_total_5,gvl_4_total_6,])

header_1 = ["Description", "Part Number", "Week 1", "Week 2", "Week 3", "Week 4", "Week 5", "Week 6"]

wk_1_ped_total = su_lopro_1[2]+su_lopro_2[2]+su_lopro_25[2]+miller_0[2]+miller_1[2]
wk_2_ped_total = su_lopro_1[3]+su_lopro_2[3]+su_lopro_25[3]+miller_0[3]+miller_1[3]
wk_3_ped_total = su_lopro_1[4]+su_lopro_2[4]+su_lopro_25[4]+miller_0[4]+miller_1[4]
wk_4_ped_total = su_lopro_1[5]+su_lopro_2[5]+su_lopro_25[5]+miller_0[5]+miller_1[5]
wk_5_ped_total = su_lopro_1[6]+su_lopro_2[6]+su_lopro_25[6]+miller_0[6]+miller_1[6]
wk_6_ped_total = su_lopro_1[7]+su_lopro_2[7]+su_lopro_25[7]+miller_0[7]+miller_1[7]

wk_1_adult_total = su_lopro_3[2]+su_lopro_4[2]+ns_lopro_3[2]+ns_lopro_4[2]+mac_3[2]+mac_4[2]+ns_mac_3[2]+ns_mac_4[2]
wk_2_adult_total = su_lopro_3[3]+su_lopro_4[3]+ns_lopro_3[3]+ns_lopro_4[3]+mac_3[3]+mac_4[3]+ns_mac_3[3]+ns_mac_4[3]
wk_3_adult_total = su_lopro_3[4]+su_lopro_4[4]+ns_lopro_3[4]+ns_lopro_4[4]+mac_3[4]+mac_4[4]+ns_mac_3[4]+ns_mac_4[4]
wk_4_adult_total = su_lopro_3[5]+su_lopro_4[5]+ns_lopro_3[5]+ns_lopro_4[5]+mac_3[5]+mac_4[5]+ns_mac_3[5]+ns_mac_4[5]
wk_5_adult_total = su_lopro_3[6]+su_lopro_4[6]+ns_lopro_3[6]+ns_lopro_4[6]+mac_3[6]+mac_4[6]+ns_mac_3[6]+ns_mac_4[6]
wk_6_adult_total = su_lopro_3[7]+su_lopro_4[7]+ns_lopro_3[7]+ns_lopro_4[7]+mac_3[7]+mac_4[7]+ns_mac_3[7]+ns_mac_4[7]

wk_1_gvl_adu_total = gvl_3[2] + gvl_4[2]
wk_2_gvl_adu_total = gvl_3[3] + gvl_4[3]
wk_3_gvl_adu_total = gvl_3[4] + gvl_4[4]
wk_4_gvl_adu_total = gvl_3[5] + gvl_4[5]
wk_5_gvl_adu_total = gvl_3[6] + gvl_4[6]
wk_6_gvl_adu_total = gvl_3[7] + gvl_4[7]

#
#REACTIVATE ALL WHEN IF'S ARE IN - convert to append to sheets when all are mostly done, who knows. what ever. no one will ever read this.
#
print(header_1)
print(su_lopro_1)
print(su_lopro_2)
print(su_lopro_25)
print(miller_0)
print(miller_1)
total_ped_su_del = ["Total Ped SU Deliveries", "", wk_1_ped_total, wk_2_ped_total, wk_3_ped_total, wk_4_ped_total, wk_5_ped_total, wk_6_ped_total]
print(total_ped_su_del)
print(su_lopro_3)
print(su_lopro_4)
print(ns_lopro_3)
print(ns_lopro_4)
print(mac_3)
print(ns_mac_3)
print(mac_4)
print(ns_mac_4)
total_adult_su_del = ["Total Adult SU Deliveries", "", wk_1_adult_total, wk_2_adult_total, wk_3_adult_total, wk_4_adult_total, wk_5_adult_total, wk_6_adult_total]
print(total_adult_su_del)
print(gvl_3)
print(gvl_4)
total_adult_gvl_su_del = ["Total Adult Stat Deliveries", "", wk_1_gvl_adu_total, wk_2_gvl_adu_total, wk_3_gvl_adu_total, wk_4_gvl_adu_total, wk_5_gvl_adu_total, wk_6_gvl_adu_total, ]
print(total_adult_gvl_su_del)




